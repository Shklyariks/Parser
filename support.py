import openpyxl
import datetime


def menu():
    print('Начать сбор информации: 1\n'
          'Проверить наличие файла с записями: 2\n'
          'Проверить обновления: 3\n'
          'Exit: 4')


def search():
    try:
        book = openpyxl.load_workbook('Banki.ru.xlsx')
        sheet = book.active
        count = sheet.max_row
        print(f'Файл найден\n'
              f'Записано строк {count}')
    except FileNotFoundError:
        print('Файл не найден')

def report():

    book = openpyxl.load_workbook('Banki.ru.xlsx')
    sheet = book.active.values
    message_list = []
    past = datetime.datetime.now() - datetime.timedelta(days=7)

    for i in sheet:
        date = datetime.datetime.strptime(i[3], ' %d.%m.%Y %H:%M')
        if date >= past:
            message_list.append(i[0])
    elements = len(message_list)
    print(f'{elements} обновлений')

    message = '\n'.join(message_list)

    return message

