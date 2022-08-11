from pprint import pprint
import openpyxl as openpyxl


def create_post(array):

    book = openpyxl.load_workbook('Banki.ru.xlsx')
    # book = openpyxl.load_workbook('banki.xlsx')
    sheet = book.active
    row = 1
    for i in array:
        sheet.cell(row=row, column=1)
        sheet.append(i)
        book.save('Banki.ru.xlsx')
    pprint('Запись завершена')